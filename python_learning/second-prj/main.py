import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}  # empty dictionary

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4)
